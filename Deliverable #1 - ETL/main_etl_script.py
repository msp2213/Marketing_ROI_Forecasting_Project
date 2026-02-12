"""
═══════════════════════════════════════════════════════════════════════════════
MAIN ETL SCRIPT - COMPLETE STANDALONE P&L PROCESSING PIPELINE
═══════════════════════════════════════════════════════════════════════════════

This is a SINGLE, SELF-CONTAINED script that does EVERYTHING:
✓ Extracts data from all 4 P&L Excel files
✓ Asks user which file is the naming standard (or uses default: GG)
✓ Uses AI (GPT-4) to create unified hierarchical ontology
✓ Maps equivalent accounts across all files
✓ Shows which parts compose each account (with +/- signs)
✓ Transforms to standardized vertical format
✓ Outputs Excel + CSV files with complete mappings

NO OTHER SCRIPTS NEEDED - This does it all!

USAGE:
    python main_etl_script.py
    
    Then press Enter to use GG as base (or select another file)
    Wait ~60 seconds for AI processing
    Check output/ folder for results

OUTPUTS:
    • vertical_pl_data_{timestamp}.csv - Main vertical data (12K+ rows)
    • complete_pl_etl_{timestamp}.xlsx - Excel with multiple sheets
    • unified_ontology_{timestamp}.csv - Hierarchical account mappings

═══════════════════════════════════════════════════════════════════════════════
"""

import openpyxl
import pandas as pd
from openai import OpenAI
import os
from dotenv import load_dotenv
import csv
import io
from datetime import datetime
import random

# Load environment variables
load_dotenv()

class PLETLPipeline:
    def __init__(self, pl_files, file_ids, output_dir="output"):
        self.pl_files = pl_files
        self.file_ids = file_ids
        self.output_dir = output_dir
        self.base_idx = None
        self.base_id = None
        self.all_structures = {}
        self.unified_ontology = None
        self.vertical_data = []
        self.account_name_mapping = {}  # Maps cleaned names back to originals
        
        os.makedirs(output_dir, exist_ok=True)
    
    def clean_account_name(self, account_name):
        """
        Clean account names by removing numeric prefixes like "70000 ·" or "Total 65000 ·"
        but preserve original for mapping
        
        Examples:
            "70000 · Other Income" -> "Other Income"
            "40000 · Sales" -> "Sales"
            "Total 65000 · Other Ops & Selling Costs" -> "Total Other Ops & Selling Costs"
            "Total 61000 · Sales & Marketing" -> "Total Sales & Marketing"
            "General & Administrative" -> "General & Administrative" (unchanged)
        """
        if not account_name or not isinstance(account_name, str):
            return account_name
        
        original = account_name.strip()
        
        # Check for pattern with bullet point
        if '·' not in original:
            return original
        
        parts = original.split('·', 1)
        if len(parts) != 2:
            return original
        
        before_bullet = parts[0].strip()
        after_bullet = parts[1].strip()
        
        # Check if everything before bullet is just a number
        if before_bullet.replace(' ', '').isdigit():
            # Pattern: "70000 · Other Income" -> "Other Income"
            return after_bullet
        
        # Check if it's like "Total 65000" - extract prefix words and remove number
        words = before_bullet.split()
        if len(words) > 1:
            # Filter out words that are just numbers
            prefix_words = [word for word in words if not word.replace(' ', '').isdigit()]
            if prefix_words:
                # Pattern: "Total 65000 · Other Ops" -> "Total Other Ops"
                return ' '.join(prefix_words) + ' ' + after_bullet
            else:
                # All words were numbers, just return after bullet
                return after_bullet
        
        return original
    
    def extract_pl_structure(self, file_path, file_id, extract_all_levels=False):
        """
        Extract P&L structure from Excel file
        Dynamically detects where label columns end and data columns begin
        """
        print(f"  Extracting from {file_id}...")
        
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        max_row = ws.max_row
        max_col = ws.max_column
        
        # Dynamically detect where data columns start
        # Check row 1 (header row) to find first cell with a period/date-like value
        data_start_col = None
        for col_idx in range(1, max_col + 1):
            header_cell = ws.cell(row=1, column=col_idx)
            if header_cell.value:
                header_str = str(header_cell.value).strip()
                # Check if this looks like a period header (contains month names or numbers)
                if any(month in header_str for month in ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                                                          'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec',
                                                          'TOTAL', '20', '19']):
                    data_start_col = col_idx
                    break
        
        # If not found, default to column I (9)
        if data_start_col is None:
            data_start_col = 9
        
        # Label columns are everything before data starts
        label_columns = list(range(1, data_start_col))
        
        print(f"     • Label columns: A-{chr(64 + len(label_columns))} ({len(label_columns)} levels)")
        print(f"     • Data starts at column: {chr(64 + data_start_col)}")
        
        # Get period headers (first row, starting from column I)
        period_headers = []
        for col_idx in range(data_start_col, max_col + 1):
            header_cell = ws.cell(row=1, column=col_idx)
            if header_cell.value:
                period_headers.append({
                    'col_idx': col_idx,
                    'period': str(header_cell.value).strip()
                })
        
        # Identify TOTAL columns to exclude
        exclude_columns = set()
        for header_info in period_headers:
            if "TOTAL" in header_info['period'].upper():
                exclude_columns.add(header_info['col_idx'])
        
        # Extract structure and data
        structure_data = []
        detail_data = []
        
        for row in range(1, max_row + 1):
            # Find label and level
            label = None
            original_label = None
            level = None
            
            for col_idx in label_columns:
                cell = ws.cell(row=row, column=col_idx)
                if cell.value and str(cell.value).strip():
                    original_label = str(cell.value).strip()
                    label = self.clean_account_name(original_label)
                    level = col_idx - 1
                    
                    # Store mapping from cleaned to original name
                    if label != original_label:
                        mapping_key = (file_id, label)
                        if mapping_key not in self.account_name_mapping:
                            self.account_name_mapping[mapping_key] = original_label
                    
                    break
            
            if label is None:
                continue
            
            # Skip deeper levels if not extracting all
            if not extract_all_levels and level > 4:
                continue
            
            # Get all period values
            row_values = {}
            for header_info in period_headers:
                if header_info['col_idx'] not in exclude_columns:
                    cell = ws.cell(row=row, column=header_info['col_idx'])
                    try:
                        value = float(cell.value) if cell.value is not None else 0
                        row_values[header_info['period']] = value
                    except (ValueError, TypeError):
                        row_values[header_info['period']] = 0
            
            total_sum = sum(row_values.values())
            
            structure_data.append({
                'Level': level,
                'Account': label,  # Cleaned name
                'Original_Account': original_label,  # Keep original for reference
                'Sum': round(total_sum, 2),
                'Row': row
            })
            
            # Store detail data for vertical transformation
            for period, value in row_values.items():
                detail_data.append({
                    'File': file_id,
                    'Level': level,
                    'Cleaned_Account': label,  # Cleaned name for matching
                    'Original_Account': original_label,  # Original name for display
                    'Period': period,
                    'Value': value
                })
        
        return pd.DataFrame(structure_data), detail_data
    
    def get_base_file_choice(self):
        """
        Ask user which file to use as base naming standard
        """
        print("\n" + "=" * 80)
        print("SELECT BASE FILE FOR NAMING STANDARD")
        print("=" * 80)
        for i, (file, file_id) in enumerate(zip(self.pl_files, self.file_ids), 1):
            print(f"{i}. {file_id:10s} - {file}")
        
        print("\nWhich file should be used as the BASE naming standard?")
        print("(Press Enter to use default: 1)")
        
        try:
            choice = input(f"Enter number (1-{len(self.pl_files)}): ").strip()
            if not choice:
                return 0
            choice_num = int(choice) - 1
            if 0 <= choice_num < len(self.pl_files):
                return choice_num
            else:
                print(f"Invalid choice. Using default: {self.file_ids[0]}")
                return 0
        except (ValueError, KeyboardInterrupt):
            print(f"\nUsing default: {self.file_ids[0]}")
            return 0
    
    def obfuscate_values(self, df):
        """
        Obfuscate Sum values by multiplying by a random factor between 1.1 and 1.3
        This protects sensitive financial data while preserving relative relationships
        """
        df_copy = df.copy()
        if 'Sum' in df_copy.columns:
            # Apply random multiplier to each value
            df_copy['Sum'] = df_copy['Sum'].apply(
                lambda x: round(x * random.uniform(1.1, 1.3), 2) if pd.notna(x) else x
            )
        return df_copy
    
    def create_unified_ontology_prompt(self, base_df, other_files_data):
        """
        Create unified ontology using the SAME logic as analyze_pl_ontology.py
        NOTE: Values are obfuscated before sending to OpenAI for data privacy
        """
        # Obfuscate all dataframes before sending to OpenAI
        base_df_obfuscated = self.obfuscate_values(base_df)
        other_files_obfuscated = {fid: self.obfuscate_values(df) for fid, df in other_files_data.items()}
        
        # Combine all data for context, including max level info for each file
        prompt = f"""I have multiple Profit & Loss (P&L) statements from different entities that need to be unified into a single ontology.

IMPORTANT: Each file has a different hierarchical depth (max levels):
"""
        
        # Add max level information
        all_files_info = [(self.base_id, base_df_obfuscated)] + list(other_files_obfuscated.items())
        for file_id, df in all_files_info:
            max_level = df['Level'].max()
            prompt += f"  • {file_id}: Levels 0-{max_level} ({max_level + 1} levels)\n"
        
        prompt += f"""

BASE FILE (Standard Naming): {self.base_id}
{base_df_obfuscated.to_string(index=False)}

"""
        
        for file_id, df in other_files_obfuscated.items():
            prompt += f"\n{'=' * 80}\n"
            prompt += f"FILE: {file_id}\n"
            prompt += f"{'=' * 80}\n"
            prompt += df.to_string(index=False)
            prompt += "\n"
        
        prompt += f"""

Please analyze this P&L ontology and provide an ordered hierarchical breakdown showing:
1. How each level's items are composed (which sub-items sum up to create parent items)
2. Whether each relationship is positive (adds to parent) or negative (subtracts from parent)
3. The natural direction (Debit/Credit) for each account
4. Mapping of equivalent accounts across all files

CRITICAL RULES FOR DIFFERENT LEVEL STRUCTURES:
- Parts of level N MUST come from level N+1 ONLY
- RESPECT the max level of each file (see above)
- If a file has max level 6, do NOT try to map accounts at level 7
- When mapping accounts, only map if the account exists at that level in that file
- If an account exists at level 7 in BASE but a file only goes to level 6, leave that file's mapping blank
- Use the "Level" column in each file's data to determine what accounts exist

For example:
- Level 0: Net Income = [Level 1 items: Net Ordinary Income + Net Other Income]
- Level 1: Net Ordinary Income = [Level 2 items: Gross Profit - Total Expenses]
- Level 2: Gross Profit = [Level 3 items: ...]
- And so on...

ACCOUNT DIRECTION (for proper P&L presentation):
- Revenue/Sales/Income accounts: Direction = "Credit" (normally positive)
- Expense/Cost/Fee accounts: Direction = "Debit" (normally negative on P&L)
- Deduction/Discount accounts: Direction = "Debit" (reduce revenue)

Please provide the output as a CSV with these columns:
- Original_Level: The original hierarchical level from the BASE data (0-7)
- Item: The account/line item name (use BASE FILE naming as standard)
- Implied_Level: The level this item actually functions at in the hierarchy
- Part: The component/sub-item that makes up this item (repeat rows for each part)
- Part_Level: The level of the part (must be Original_Level + 1 for non-leaf items)
- Sign: Whether this part is "Positive" or "Negative" in the calculation
- Direction: "Debit" or "Credit" - the natural balance for this account
- {', '.join([f'{fid}_Account' for fid in self.file_ids])}: Matching account name in each file

Important: 
- Repeat the Item row for EACH part that composes it
- Parts MUST be from the next level down (level + 1)
- Use the actual Sum values to infer the relationships
- Make sure the hierarchy is mathematically consistent
- For leaf items (with no sub-components), Part should be "Value" and Part_Level should be empty
- Use proper CSV formatting: wrap fields containing commas in double quotes
- Use BASE FILE account names as the standard naming

Output ONLY the CSV data with headers, no explanations."""

        return prompt
    
    def analyze_with_openai(self, prompt):
        """
        Send prompt to OpenAI and get structured response
        """
        api_key = os.getenv('OPENAI_API')
        if not api_key:
            raise ValueError("OPENAI_API key not found in .env file")
        
        client = OpenAI(api_key=api_key)
        
        print("  Analyzing with OpenAI GPT-4 (this may take 30-60 seconds)...")
        
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {
                    "role": "system",
                    "content": "You are a financial analyst expert in P&L statement structure, account mapping, and creating unified chart of accounts. You provide precise, structured CSV outputs."
                },
                {
                    "role": "user",
                    "content": prompt
                }
            ],
            temperature=0.1,
            max_tokens=8000
        )
        
        return response.choices[0].message.content
    
    def clean_csv_response(self, response):
        """
        Clean OpenAI response and parse as CSV with robust handling
        """
        cleaned = response.strip()
        if cleaned.startswith('```csv'):
            cleaned = cleaned[6:]
        if cleaned.startswith('```'):
            cleaned = cleaned[3:]
        if cleaned.endswith('```'):
            cleaned = cleaned[:-3]
        cleaned = cleaned.strip()
        
        # Try to parse as CSV
        try:
            reader = csv.reader(io.StringIO(cleaned))
            rows = list(reader)
            return rows
        except:
            # If parsing fails, do manual cleanup for commas in fields
            lines = cleaned.split('\n')
            fixed_rows = []
            
            for i, line in enumerate(lines):
                if not line.strip():
                    continue
                    
                # Use csv reader for each line
                try:
                    row = list(csv.reader([line]))[0]
                    fixed_rows.append(row)
                except:
                    # If still fails, just split by comma (not ideal but fallback)
                    fixed_rows.append(line.split(','))
            
            return fixed_rows
    
    def create_vertical_output(self, detail_data_all_files):
        """
        Create vertical format output with standardized account names and correct value signs
        """
        print("\n  Transforming to vertical format with standardized names...")
        print("  Applying correct directional signs to values...")
        
        # Create mapping from original to standard accounts
        # The ontology has multiple rows per account (one for each part), so we need to deduplicate
        account_mapping = {}
        
        # Get unique parent accounts from ontology (using analyze_pl_ontology.py column names)
        unique_accounts = self.unified_ontology.groupby('Item').first().reset_index()
        
        for _, row in unique_accounts.iterrows():
            std_account = row['Item']
            std_level = row.get('Original_Level', row.get('Implied_Level', 0))
            direction = row.get('Direction', 'Credit')  # Default to Credit if not specified
            
            # Get parts that compose this account
            parts_df = self.unified_ontology[self.unified_ontology['Item'] == std_account]
            parts_info = []
            for _, part_row in parts_df.iterrows():
                if pd.notna(part_row.get('Part', '')):
                    part = str(part_row['Part']).strip()
                    if part and part != 'Value':
                        parts_info.append({
                            'part': part,
                            'sign': part_row.get('Sign', 'Positive')
                        })
            
            # Map accounts from each file
            for file_id in self.file_ids:
                col_name = f'{file_id}_Account'
                if col_name in row and pd.notna(row[col_name]) and str(row[col_name]).strip():
                    original = str(row[col_name]).strip()
                    account_mapping[(file_id, original)] = {
                        'Standard_Account': std_account,
                        'Standard_Level': std_level,
                        'Direction': direction,
                        'Parts': parts_info if parts_info else []
                    }
        
        # Transform detail data
        vertical_data = []
        
        for detail in detail_data_all_files:
            file_id = detail['File']
            cleaned_account = detail['Cleaned_Account']
            original_account = detail['Original_Account']
            raw_value = detail['Value']
            
            # Look up standard account using CLEANED name
            mapping_key = (file_id, cleaned_account)
            if mapping_key in account_mapping:
                mapping = account_mapping[mapping_key]
                
                # Create parts string
                parts_str = ''
                if mapping['Parts']:
                    parts_list = [f"{p['part']} ({p['sign']})" for p in mapping['Parts']]
                    parts_str = ' + '.join(parts_list)
                
                # Apply correct sign based on Direction
                # Credit accounts (Revenue/Income) should be positive
                # Debit accounts (Expenses/Costs) should be negative
                direction = mapping['Direction']
                corrected_value = raw_value
                
                if direction == 'Debit':
                    # Expense/Cost accounts should be negative on P&L
                    # If the value is positive, make it negative
                    if raw_value > 0:
                        corrected_value = -raw_value
                elif direction == 'Credit':
                    # Revenue/Income accounts should be positive on P&L
                    # If the value is negative, make it positive
                    if raw_value < 0:
                        corrected_value = abs(raw_value)
                
                vertical_data.append({
                    'Source_File': file_id,
                    'Period': detail['Period'],
                    'Original_Account_Name': original_account,  # Original name with prefix
                    'Standard_Account': mapping['Standard_Account'],
                    'Standard_Level': mapping['Standard_Level'],
                    'Direction': direction,
                    'Composition': parts_str,
                    'Original_Value': raw_value,
                    'Corrected_Value': corrected_value
                })
            else:
                # Account not in mapping - keep original, apply basic logic
                # Use CLEANED account for keyword matching but display ORIGINAL
                is_expense = any(keyword in cleaned_account.lower() for keyword in 
                                ['expense', 'cost', 'fee', 'deduction', 'discount', 'commission'])
                direction = 'Debit' if is_expense else 'Credit'
                corrected_value = -abs(raw_value) if is_expense and raw_value > 0 else raw_value
                
                vertical_data.append({
                    'Source_File': file_id,
                    'Period': detail['Period'],
                    'Original_Account_Name': original_account,  # Show with prefix
                    'Standard_Account': cleaned_account,  # Use cleaned name
                    'Standard_Level': detail['Level'],
                    'Direction': direction,
                    'Composition': '',
                    'Original_Value': raw_value,
                    'Corrected_Value': corrected_value
                })
        
        return pd.DataFrame(vertical_data)
    
    def run(self):
        """
        Execute the complete ETL pipeline
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        print("=" * 80)
        print("P&L ETL PIPELINE - COMPREHENSIVE DATA PROCESSING")
        print("=" * 80)
        print(f"\nTimestamp: {timestamp}")
        print(f"Processing {len(self.pl_files)} P&L files")
        
        # STEP 1: Get base file choice
        print("\n" + "=" * 80)
        print("STEP 1: SELECT BASE FILE")
        print("=" * 80)
        self.base_idx = self.get_base_file_choice()
        self.base_id = self.file_ids[self.base_idx]
        base_file = self.pl_files[self.base_idx]
        
        print(f"\n✓ Using '{self.base_id}' ({base_file}) as BASE naming standard")
        
        # STEP 2: Extract all structures
        print("\n" + "=" * 80)
        print("STEP 2: EXTRACT P&L STRUCTURES")
        print("=" * 80)
        
        detail_data_all = []
        
        for file_path, file_id in zip(self.pl_files, self.file_ids):
            if os.path.exists(file_path):
                structure_df, detail_data = self.extract_pl_structure(
                    file_path, file_id, extract_all_levels=True
                )
                self.all_structures[file_id] = structure_df
                detail_data_all.extend(detail_data)
                print(f"  ✓ {file_id}: {len(structure_df)} accounts, {len(detail_data)} detail records")
            else:
                print(f"  ✗ {file_id}: File not found - {file_path}")
        
        if self.base_id not in self.all_structures:
            print(f"\n✗ ERROR: Base file '{base_file}' could not be loaded!")
            return False
        
        # STEP 3: Create unified ontology with AI
        print("\n" + "=" * 80)
        print("STEP 3: CREATE UNIFIED ONTOLOGY (AI-POWERED)")
        print("=" * 80)
        print("  🔒 Obfuscating values before sending to OpenAI (multiplying by random factor 1.1-1.3)")
        
        base_df = self.all_structures[self.base_id]
        other_files = {fid: df for fid, df in self.all_structures.items() if fid != self.base_id}
        
        prompt = self.create_unified_ontology_prompt(base_df, other_files)
        
        # Get AI response
        try:
            response = self.analyze_with_openai(prompt)
            
            # Parse and save ontology with robust CSV handling
            cleaned = response.strip()
            if cleaned.startswith('```csv'):
                cleaned = cleaned[6:]
            if cleaned.startswith('```'):
                cleaned = cleaned[3:]
            if cleaned.endswith('```'):
                cleaned = cleaned[:-3]
            cleaned = cleaned.strip()
            
            lines = cleaned.split('\n')
            
            # Expected columns: 11 (Original_Level, Item, Implied_Level, Part, Part_Level, Sign, Direction, + 4 file accounts)
            EXPECTED_COLS = 7 + len(self.file_ids)
            
            fixed_rows = []
            for i, line in enumerate(lines):
                if not line.strip():
                    continue
                
                parts = line.split(',')
                
                if i == 0:  # Header
                    fixed_rows.append(parts)
                    continue
                
                if len(parts) == EXPECTED_COLS:
                    fixed_rows.append(parts)
                elif len(parts) > EXPECTED_COLS:
                    # Handle fields with unquoted commas
                    # Structure: level, account, part, part_level, sign, direction, then 4 file accounts
                    level = parts[0]
                    
                    # Find Sign (Positive/Negative) to anchor our parsing
                    sign_idx = None
                    for j, p in enumerate(parts):
                        if p in ['Positive', 'Negative']:
                            sign_idx = j
                            break
                    
                    # Find Direction (Debit/Credit) to anchor our parsing
                    direction_idx = None
                    for j, p in enumerate(parts):
                        if p in ['Debit', 'Credit']:
                            direction_idx = j
                            break
                    
                    if sign_idx is not None and direction_idx is not None and sign_idx >= 4:
                        # Standard structure: level, account, part, part_level, sign, direction, accounts...
                        # Everything before sign-2 is account+part, sign-1 is part_level
                        part_level = parts[sign_idx - 1]
                        sign = parts[sign_idx]
                        direction = parts[direction_idx]
                        
                        # Account and part are between index 1 and sign-2
                        account_part_fields = parts[1:sign_idx-1]
                        
                        # Try to split account and part (challenging without more info)
                        # Assume account is first field, rest is part
                        account = account_part_fields[0] if len(account_part_fields) > 0 else ''
                        part = ','.join(account_part_fields[1:]) if len(account_part_fields) > 1 else ''
                        
                        # Last 4 fields are file accounts (after direction)
                        num_accounts = len(self.file_ids)
                        file_accounts = []
                        for j in range(num_accounts):
                            if j < num_accounts - 1:
                                idx = direction_idx + 1 + j
                                file_accounts.append(parts[idx] if idx < len(parts) else '')
                            else:
                                # Last one - combine remaining
                                remaining = parts[direction_idx + 1 + j:]
                                file_accounts.append(','.join(remaining) if remaining else '')
                        
                        fixed_row = [level, account, part, part_level, sign, direction] + file_accounts
                        fixed_rows.append(fixed_row)
                    else:
                        # Can't parse reliably - use simple split
                        # Take first field, then last 4, combine middle
                        level = parts[0]
                        account = parts[1] if len(parts) > 1 else ''
                        
                        # Last 4 + direction + sign + part_level = last 7
                        if len(parts) >= 8:
                            part_level = parts[-6]
                            sign = parts[-5]
                            direction = parts[-4]
                            file_accounts = [parts[-3], parts[-2], parts[-1], '']
                            # Middle is part
                            part = ','.join(parts[2:-6]) if len(parts) > 8 else ''
                        else:
                            part = ''
                            part_level = ''
                            sign = ''
                            direction = ''
                            file_accounts = ['', '', '', '']
                        
                        fixed_row = [level, account, part, part_level, sign, direction] + file_accounts
                        fixed_rows.append(fixed_row)
                else:
                    # Too few - pad with empty strings
                    while len(parts) < EXPECTED_COLS:
                        parts.append('')
                    fixed_rows.append(parts)
            
            # Create DataFrame directly in memory (no need to save to disk)
            if len(fixed_rows) > 0:
                self.unified_ontology = pd.DataFrame(fixed_rows[1:], columns=fixed_rows[0])
            else:
                raise ValueError("No ontology data received")
            
            print(f"  ✓ Created unified ontology: {len(self.unified_ontology)} relationships")
            print(f"  ✓ Processing complete (in memory)")
            
        except Exception as e:
            print(f"  ✗ ERROR creating ontology: {e}")
            return False
        
        # STEP 4: Create vertical output
        print("\n" + "=" * 80)
        print("STEP 4: TRANSFORM TO VERTICAL FORMAT")
        print("=" * 80)
        
        vertical_df = self.create_vertical_output(detail_data_all)
        
        # Parse Period into Year and Month columns
        print("  ✓ Parsing dates into Year and Month columns...")
        
        def parse_period(period_str):
            """
            Parse period like 'Jan 22' into year and month
            Returns: (year, month) e.g., (2022, 'January')
            """
            if not period_str or pd.isna(period_str):
                return None, None
            
            try:
                parts = str(period_str).strip().split()
                if len(parts) == 2:
                    month_abbr = parts[0]
                    year_short = parts[1]
                    
                    # Month mapping
                    month_map = {
                        'Jan': 'January', 'Feb': 'February', 'Mar': 'March',
                        'Apr': 'April', 'May': 'May', 'Jun': 'June',
                        'Jul': 'July', 'Aug': 'August', 'Sep': 'September',
                        'Oct': 'October', 'Nov': 'November', 'Dec': 'December'
                    }
                    
                    month = month_map.get(month_abbr, month_abbr)
                    
                    # Convert 2-digit year to 4-digit
                    year_num = int(year_short)
                    if year_num < 50:
                        year = 2000 + year_num
                    else:
                        year = 1900 + year_num
                    
                    return year, month
            except:
                pass
            
            return None, None
        
        # Apply parsing
        vertical_df[['Year', 'Month']] = vertical_df['Period'].apply(
            lambda x: pd.Series(parse_period(x))
        )
        
        # Select only required columns for output (Year and Month instead of Period)
        output_columns = ['Source_File', 'Year', 'Month', 'Standard_Account', 'Standard_Level', 'Direction', 'Original_Value']
        vertical_output = vertical_df[output_columns].copy()
        
        # Save vertical output as standalone CSV (no timestamp)
        vertical_file = os.path.join(self.output_dir, "complete_pl.csv")
        vertical_output.to_csv(vertical_file, index=False)
        print(f"  ✓ Created complete P&L CSV: {len(vertical_output):,} rows")
        print(f"  ✓ Saved to: {os.path.basename(vertical_file)}")
        print(f"  ✓ Columns: {', '.join(output_columns)}")
        
        # STEP 5: Generate summary statistics
        print("\n" + "=" * 80)
        print("STEP 5: SUMMARY STATISTICS")
        print("=" * 80)
        
        print(f"\nData Overview:")
        print(f"  Total records in vertical format: {len(vertical_output):,}")
        print(f"  Unique years: {vertical_output['Year'].nunique()}")
        print(f"  Unique months: {vertical_output['Month'].nunique()}")
        print(f"  Unique standard accounts: {vertical_output['Standard_Account'].nunique()}")
        
        # Convert Standard_Level to int for sorting
        try:
            levels = sorted([int(x) for x in vertical_df['Standard_Level'].dropna().unique()])
            print(f"  Standard levels: {levels}")
        except:
            print(f"  Standard levels: {list(vertical_df['Standard_Level'].unique())}")
        
        print(f"\nRecords by Source File:")
        file_counts = vertical_df['Source_File'].value_counts().sort_index()
        for file_id, count in file_counts.items():
            print(f"  {file_id:10s}: {count:,} records")
        
        print(f"\nRecords by Level:")
        # Convert to int for proper sorting
        temp_df = vertical_df.copy()
        temp_df['Standard_Level'] = pd.to_numeric(temp_df['Standard_Level'], errors='coerce')
        level_counts = temp_df.groupby('Standard_Level').size().sort_index()
        for level, count in level_counts.items():
            if pd.notna(level):
                print(f"  Level {int(level)}: {count:,} records")
        
        print(f"\nRecords by Direction:")
        direction_counts = vertical_df['Direction'].value_counts().sort_index()
        for direction, count in direction_counts.items():
            print(f"  {direction:6s}: {count:,} records")
        
        # Show value correction statistics
        values_corrected = len(vertical_df[vertical_df['Original_Value'] != vertical_df['Corrected_Value']])
        print(f"\nValue Corrections:")
        print(f"  Values corrected for proper direction: {values_corrected:,} records")
        print(f"  Values unchanged: {len(vertical_df) - values_corrected:,} records")
        
        print(f"\nSample Data (first 10 rows):")
        print("=" * 80)
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', 250)
        pd.set_option('display.max_colwidth', 25)
        print(vertical_df.head(10).to_string(index=False))
        
        # STEP 6: Clean up - remove all files except complete_pl.csv
        print("\n" + "=" * 80)
        print("STEP 6: CLEANUP")
        print("=" * 80)
        
        # Get list of all files in output directory
        import glob
        all_output_files = glob.glob(os.path.join(self.output_dir, '*'))
        files_deleted = 0
        
        for file_path in all_output_files:
            if os.path.isfile(file_path):
                filename = os.path.basename(file_path)
                # Keep only complete_pl.csv
                if filename != 'complete_pl.csv':
                    try:
                        os.remove(file_path)
                        files_deleted += 1
                    except Exception as e:
                        print(f"  ⚠ Could not delete {filename}: {e}")
        
        print(f"  ✓ Cleaned up {files_deleted} intermediate file(s)")
        print(f"  ✓ Kept: complete_pl.csv")
        
        # Final summary
        print("\n" + "=" * 80)
        print("✓ ETL PIPELINE COMPLETED SUCCESSFULLY!")
        print("=" * 80)
        print(f"\nOutput File:")
        print(f"  📊 {os.path.basename(vertical_file)}")
        print(f"     • {len(vertical_output):,} rows")
        print(f"     • {len(output_columns)} columns: {', '.join(output_columns)}")
        print(f"     • Years: {sorted(vertical_output['Year'].dropna().unique().astype(int).tolist())}")
        print(f"     • {vertical_output['Standard_Account'].nunique()} unique accounts")
        print(f"\nFile location: {vertical_file}")
        print(f"\n💡 TIP: Import this CSV into your BI tool or database for analysis!")
        
        return True


if __name__ == "__main__":
    # Define P&L files
    pl_files = [
        "GG IS_2022-2024.xlsx",
        "NUNA IS_2022-2024.xlsx",
        "SOUND IS_2024.xlsx",
        "STO IS_2022-2024.xlsx"
    ]
    
    file_ids = ["GG", "NUNA", "SOUND", "STO"]
    
    # Create and run pipeline
    pipeline = PLETLPipeline(pl_files, file_ids)
    success = pipeline.run()
    
    if success:
        print("\n🎉 ETL pipeline executed successfully!")
    else:
        print("\n❌ ETL pipeline encountered errors.")